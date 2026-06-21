#!/usr/bin/env python3
"""
Import game library from a simplified Excel file.

Expected Excel format (single sheet, 3 columns):
    | Game Name              | Storefront | Gamer ID          |
    |------------------------|------------|-------------------|
    | Baldur's Gate 3        | Steam      | jijo_george_max   |
    | Cyberpunk 2077         | Epic       | Geekstradamus01   |
    | The Witcher 3          | GOG        | jijo_george       |

A "GeForce NOW Catalog" sheet is also supported (optional):
    | No. | Title | Publisher | Available Store(s) |

Usage:
    python scripts/import_games_from_excel.py <path_to_excel_file>
"""

import sys
import os
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app, db
from app.models import (
    Platform, Account, GameMaster,
    OwnedGameSteam, OwnedGameEpic, OwnedGameGog, 
    OwnedGameEA, OwnedGameBattlenet, OwnedGameUbisoft,
    GfnGame,
)


STOREFRONT_MAP = {
    "steam": "Steam",
    "epic": "Epic",
    "epic games": "Epic",
    "epic games store": "Epic",
    "gog": "GOG",
    "gog.com": "GOG",
    "ea": "EA",
    "ea play": "EA",
    "origin": "EA",
    "battle.net": "Battle.net",
    "battlenet": "Battle.net",
    "blizzard": "Battle.net",
    "ubisoft": "Ubisoft",
    "ubisoft connect": "Ubisoft",
    "uplay": "Ubisoft",
}


class GameImporter:
    """Import games from simplified Excel."""

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = openpyxl.load_workbook(excel_path)
        self.app = create_app()
        self.app_context = self.app.app_context()
        self.app_context.push()
        self.stats = {"created": 0, "skipped": 0, "errors": 0, "gfn": 0}

    def log(self, msg, level="INFO"):
        print(f"[{level}] {msg}")

    def is_valid(self, title):
        if not title or not isinstance(title, str):
            return False
        t = str(title).strip()
        return len(t) >= 2 and not t.startswith("=")

    def normalise_storefront(self, raw):
        if not raw:
            return None
        return STOREFRONT_MAP.get(str(raw).strip().lower())

    def get_or_create_platform(self, name):
        p = db.session.query(Platform).filter_by(name=name).first()
        if not p:
            p = Platform(name=name)
            db.session.add(p)
            db.session.flush()
            self.log(f"Created platform: {name}")
        return p

    def get_or_create_account(self, platform, username):
        a = db.session.query(Account).filter_by(
            platform_id=platform.platform_id, username=username
        ).first()
        if not a:
            a = Account(
                platform_id=platform.platform_id,
                username=username,
                display_name=username,
            )
            db.session.add(a)
            db.session.flush()
            self.log(f"Created account: {platform.name}/{username}")
        return a

    def get_or_create_game(self, title):
        normalised = title.lower().strip()
        g = db.session.query(GameMaster).filter(
            db.func.lower(GameMaster.title) == normalised
        ).first()
        if not g:
            g = GameMaster(title=title, status="Not Started")
            db.session.add(g)
            db.session.flush()
        return g

    def add_ownership(self, game, account, storefront_name):
        """Create ownership record in the correct platform table."""
        if storefront_name == "Steam":
            existing = db.session.query(OwnedGameSteam).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameSteam(
                account_id=account.account_id,
                game_id=game.game_id,
                steam_appid=game.game_id,
                playtime_minutes=0,
            ))

        elif storefront_name == "Epic":
            existing = db.session.query(OwnedGameEpic).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameEpic(
                account_id=account.account_id,
                game_id=game.game_id,
                epic_item_id=str(game.game_id),
            ))

        elif storefront_name == "GOG":
            existing = db.session.query(OwnedGameGog).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameGog(
                account_id=account.account_id,
                game_id=game.game_id,
                gog_product_id=str(game.game_id),
            ))

        elif storefront_name == "EA":
            existing = db.session.query(OwnedGameEA).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameEA(
                account_id=account.account_id,
                game_id=game.game_id,
                ea_game_id=str(game.game_id),
            ))

        elif storefront_name == "Battle.net":
            existing = db.session.query(OwnedGameBattlenet).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameBattlenet(
                account_id=account.account_id,
                game_id=game.game_id,
                battlenet_game_id=str(game.game_id),
            ))

        elif storefront_name == "Ubisoft":
            existing = db.session.query(OwnedGameUbisoft).filter_by(
                account_id=account.account_id, game_id=game.game_id
            ).first()
            if existing:
                return False
            db.session.add(OwnedGameUbisoft(
                account_id=account.account_id,
                game_id=game.game_id,
                ubisoft_game_id=str(game.game_id),
            ))

        else:
            return False

        return True

    def detect_format(self):
        """Detect if Excel uses simplified (3-col) or legacy format."""
        ws = self.wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        # Simplified: Game Name, Storefront, Gamer ID
        if any("game" in h and "name" in h for h in headers):
            return "simplified"
        if any("storefront" in h for h in headers):
            return "simplified"

        # Legacy: multiple sheets per platform
        if any("Epic Library" in s or "Steam Library" in s or "GOG Games" in s for s in self.wb.sheetnames):
            return "legacy"

        # Default to simplified
        return "simplified"

    def import_simplified(self):
        """Import from simplified 3-column format."""
        ws = self.wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        # Find column indices
        name_col = next((i for i, h in enumerate(headers) if "game" in h or "name" in h or "title" in h), 0)
        store_col = next((i for i, h in enumerate(headers) if "store" in h or "platform" in h), 1)
        gamer_col = next((i for i, h in enumerate(headers) if "gamer" in h or "account" in h or "user" in h or "id" in h), 2)

        self.log(f"Detected columns: name={name_col}, storefront={store_col}, gamer_id={gamer_col}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                title = str(row[name_col]).strip() if row[name_col] else None
                storefront_raw = str(row[store_col]).strip() if row[store_col] else None
                gamer_id = str(row[gamer_col]).strip() if row[gamer_col] else None

                if not self.is_valid(title):
                    continue

                storefront = self.normalise_storefront(storefront_raw)
                if not storefront:
                    self.log(f"  ⚠️ Row {row_idx}: Unknown storefront '{storefront_raw}', skipping", "WARN")
                    self.stats["errors"] += 1
                    continue

                if not gamer_id:
                    self.log(f"  ⚠️ Row {row_idx}: No gamer ID, skipping", "WARN")
                    self.stats["errors"] += 1
                    continue

                # Create records
                platform = self.get_or_create_platform(storefront)
                account = self.get_or_create_account(platform, gamer_id)
                game = self.get_or_create_game(title)

                if self.add_ownership(game, account, storefront):
                    self.stats["created"] += 1
                    if row_idx <= 20 or row_idx % 100 == 0:
                        self.log(f"  ✅ [{row_idx}] {title} ({storefront}/{gamer_id})")
                else:
                    self.stats["skipped"] += 1

            except Exception as e:
                self.log(f"  ❌ Row {row_idx}: {e}", "ERROR")
                self.stats["errors"] += 1

        db.session.commit()

    def import_gfn_catalog(self):
        """Import GeForce NOW catalog sheet if present."""
        if "GeForce NOW Catalog" not in self.wb.sheetnames:
            return

        ws = self.wb["GeForce NOW Catalog"]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                title = row[1]
                if not self.is_valid(title):
                    continue
                title = str(title).strip()
                available_stores = row[3] or ""

                game = self.get_or_create_game(title)

                existing = db.session.query(GfnGame).filter_by(game_id=game.game_id).first()
                if existing:
                    continue

                stores_lower = available_stores.lower()
                gfn = GfnGame(
                    game_id=game.game_id,
                    available_on_steam="steam" in stores_lower,
                    available_on_epic="epic" in stores_lower,
                    available_on_gog="gog" in stores_lower,
                )
                db.session.add(gfn)
                self.stats["gfn"] += 1

            except Exception as e:
                self.log(f"  ❌ GFN Row {row_idx}: {e}", "ERROR")

        db.session.commit()
        self.log(f"Imported {self.stats['gfn']} GeForce NOW games")

    def run(self):
        try:
            fmt = self.detect_format()
            self.log(f"Detected format: {fmt}")
            self.log(f"Sheets: {self.wb.sheetnames}")

            if fmt == "simplified":
                self.import_simplified()
            else:
                # Legacy format — import from named sheets
                self.import_simplified()  # Still use simplified logic on active sheet

            # Always try GFN catalog
            self.import_gfn_catalog()

            db.session.commit()
            self.print_summary()

        except Exception as e:
            db.session.rollback()
            self.log(f"FATAL: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            raise
        finally:
            self.app_context.pop()

    def print_summary(self):
        print("\n" + "=" * 80)
        print("IMPORT SUMMARY")
        print("=" * 80)
        print(f"✅ Games imported:  {self.stats['created']}")
        print(f"⏭️  Already existed: {self.stats['skipped']}")
        print(f"⚠️  Errors/skipped: {self.stats['errors']}")
        print(f"☁️  GFN catalog:    {self.stats['gfn']}")
        print("=" * 80 + "\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_games_from_excel.py <path_to_excel_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    importer = GameImporter(excel_file)
    importer.run()
